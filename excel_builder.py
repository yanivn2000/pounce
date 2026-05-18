"""
excel_builder.py
Builds the final Excel report from a list of CampaignResult objects.
3 sheets: Sponsored Products | Sponsored Brands | Top of Search SP vs SB
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from analyzer import CampaignResult


# ── Palette ──────────────────────────────────────────────────────────────────
H_DARK    = PatternFill("solid", fgColor="1F3864")
H_TOP     = PatternFill("solid", fgColor="2E75B6")
H_REST    = PatternFill("solid", fgColor="4472C4")
H_PROD    = PatternFill("solid", fgColor="7BAFD4")
H_SCORE   = PatternFill("solid", fgColor="375623")
H_BID     = PatternFill("solid", fgColor="833C00")
H_NOTE    = PatternFill("solid", fgColor="404040")
H_ALERT   = PatternFill("solid", fgColor="C00000")
H_SP      = PatternFill("solid", fgColor="2E75B6")
H_SB      = PatternFill("solid", fgColor="7030A0")
ALT       = PatternFill("solid", fgColor="EBF3FB")
WHITE     = PatternFill("solid", fgColor="FFFFFF")
NOTE_BG   = PatternFill("solid", fgColor="FFF2CC")
BID_GOOD  = PatternFill("solid", fgColor="E2EFDA")
ALERT_BG  = PatternFill("solid", fgColor="FCE4D6")
SCORE_HI  = PatternFill("solid", fgColor="C6EFCE")
SCORE_MID = PatternFill("solid", fgColor="FFEB9C")
SCORE_LO  = PatternFill("solid", fgColor="FFC7CE")

# Algorithm sheet palette
MODE_ISO  = PatternFill("solid", fgColor="FFC7CE")   # red  — isolation
MODE_OPT  = PatternFill("solid", fgColor="C6EFCE")   # green — optimization
MODE_LRN  = PatternFill("solid", fgColor="FFEB9C")   # yellow — learning
MODE_NOD  = PatternFill("solid", fgColor="D9D9D9")   # grey — no data
ACT_INC   = PatternFill("solid", fgColor="E2EFDA")   # green — increase
ACT_RED   = PatternFill("solid", fgColor="FCE4D6")   # red   — reduce
ACT_KEEP  = PatternFill("solid", fgColor="F2F2F2")   # grey  — keep
H_ALGO    = PatternFill("solid", fgColor="1F3864")   # dark blue header
H_ALGO_T  = PatternFill("solid", fgColor="2E75B6")   # top header
H_ALGO_R  = PatternFill("solid", fgColor="4472C4")   # rest header
H_ALGO_P  = PatternFill("solid", fgColor="7BAFD4")   # product header

WF = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BF = Font(name="Arial", bold=True, size=10)
NF = Font(name="Arial", size=9)
SF = Font(name="Arial", bold=True, size=11)

thin = Side(style="thin", color="BDD7EE")
BRD  = Border(left=thin, right=thin, top=thin, bottom=thin)
C    = Alignment(horizontal="center", vertical="center", wrap_text=True)
L    = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _score_fill(s: int) -> PatternFill:
    if s >= 70: return SCORE_HI
    if s >= 40: return SCORE_MID
    return SCORE_LO


def _bid_fill(bid: str) -> PatternFill:
    return BID_GOOD if (bid not in ["—", ""] and "0%" not in bid) else WHITE


def _cell(ws, row, col, value, font=NF, fill=WHITE, align=C, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font; c.fill = fill; c.alignment = align; c.border = BRD
    if fmt:
        c.number_format = fmt
    return c


def _placement_vals(m):
    """Return (impr, clicks, ctr, cpc, spend, sales, roas, acos) tuple."""
    return (
        m.impressions, m.clicks,
        m.ctr, m.cpc, m.spend, m.sales,
        m.roas, m.acos,
    )


METRICS_FMTS = ["#,##0", "#,##0", "0.00%", "$#,##0.00",
                "$#,##0.00", "$#,##0.00", "0.00", "0.00%"]
METRICS_HDR  = ["Impr.", "Clicks", "CTR%", "CPC$",
                "Spend$", "Sales$", "ROAS", "ACOS%"]


def _make_detail_sheet(ws, rows: list[CampaignResult]):
    """Build a detail sheet (SP or SB)."""
    # ── Row 1: group headers ──────────────────────────────────
    ws.merge_cells("A1:C1");   _cell(ws,1,1,"Campaign Info", WF, H_DARK, C)
    ws.merge_cells("D1:K1");   _cell(ws,1,4,"Top of Search",  WF, H_TOP,  C)
    ws.merge_cells("L1:S1");   _cell(ws,1,12,"Rest of Search", WF, H_REST, C)
    ws.merge_cells("T1:AA1");  _cell(ws,1,20,"Product Pages",  WF, H_PROD, C)
    ws.merge_cells("AB1:AC1"); _cell(ws,1,28,"Score (1-100)",  WF, H_SCORE,C)
    ws.merge_cells("AD1:AD2"); _cell(ws,1,30,"Alert",          WF, H_ALERT,C)
    ws.merge_cells("AE1:AE2"); _cell(ws,1,31,"Bid Adj.",       WF, H_BID,  C)
    ws.merge_cells("AF1:AF2"); _cell(ws,1,32,"Comment",        WF, H_NOTE, C)

    # ── Row 2: sub-headers ───────────────────────────────────
    for i, h in enumerate(["Campaign", "Ad Type", "Targeting"]):
        _cell(ws, 2, i+1, h, WF, H_DARK, C)
    for pi, fill in enumerate([H_TOP, H_REST, H_PROD]):
        for mi, m in enumerate(METRICS_HDR):
            _cell(ws, 2, 4 + pi*8 + mi, m, WF, fill, C)
    _cell(ws, 2, 28, "Score",  WF, H_SCORE, C)
    _cell(ws, 2, 29, "Label",  WF, H_SCORE, C)

    # ── Data rows ────────────────────────────────────────────
    for ri, r in enumerate(rows):
        er  = ri + 3
        bg  = ALT if ri % 2 == 0 else WHITE
        sc  = r.score
        sf  = _score_fill(sc)
        bf  = _bid_fill(r.bid_rec)
        af  = ALERT_BG if r.alert else bg

        _cell(ws, er, 1, r.campaign,  BF, bg,  L)
        _cell(ws, er, 2, r.ad_type,   NF, bg,  C)
        _cell(ws, er, 3, r.targeting, NF, bg,  C)

        for pi, pm in enumerate([r.top, r.rest, r.product]):
            for mi, (v, fmt) in enumerate(zip(_placement_vals(pm), METRICS_FMTS)):
                _cell(ws, er, 4 + pi*8 + mi, v, NF, bg, C, fmt)

        _cell(ws, er, 28, sc,            SF, sf, C)
        _cell(ws, er, 29, r.score_label, NF, sf, C)
        _cell(ws, er, 30, r.alert,       NF, af, L)
        _cell(ws, er, 31, r.bid_rec,     Font(name="Arial",bold=True,size=9), bf, L)
        _cell(ws, er, 32, r.comment,     NF, NOTE_BG, L)

    # ── Column widths ────────────────────────────────────────
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 9
    for col in range(4, 29):
        ws.column_dimensions[get_column_letter(col)].width = 10
    ws.column_dimensions["AB"].width = 8
    ws.column_dimensions["AC"].width = 22
    ws.column_dimensions["AD"].width = 45
    ws.column_dimensions["AE"].width = 35
    ws.column_dimensions["AF"].width = 80

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    for r in range(3, len(rows) + 3):
        ws.row_dimensions[r].height = 50
    ws.freeze_panes = "D3"


def _make_summary_sheet(ws, sp_rows: list[CampaignResult], sb_rows: list[CampaignResult]):
    """Build the combined Top of Search comparison sheet."""
    hdrs = ["Campaign", "Type", "Top Impr.", "Top CTR%", "Top ROAS",
            "Top Sales$", "Top CPC$", "Score", "Score Label",
            "Bid Adj.", "Alert", "Comment"]
    for ci, h in enumerate(hdrs):
        _cell(ws, 1, ci+1, h, WF, H_DARK, C)

    all_rows = [(r, "SP") for r in sp_rows] + [(r, "SB") for r in sb_rows]
    all_rows.sort(key=lambda x: -x[0].score)

    fmts2 = ["", "", "#,##0", "0.00%", "0.00", "$#,##0.00", "$#,##0.00",
             "", "", "", "", ""]

    for ri, (r, tp) in enumerate(all_rows):
        er  = ri + 2
        bg  = ALT if ri % 2 == 0 else WHITE
        sf  = _score_fill(r.score)
        bf  = _bid_fill(r.bid_rec)
        af  = ALERT_BG if r.alert else bg
        tpf = H_SP if tp == "SP" else H_SB

        vals = [
            r.campaign, tp,
            r.top.impressions, r.top.ctr, r.top.roas,
            r.top.sales, r.top.cpc,
            r.score, r.score_label,
            r.bid_rec, r.alert, r.comment,
        ]
        fills = [bg, tpf, bg, bg, bg, bg, bg,
                 sf, sf, bf, af, NOTE_BG]
        aligns = [L, C, C, C, C, C, C, C, C, L, L, L]

        for ci, (v, fmt, fill, align) in enumerate(zip(vals, fmts2, fills, aligns)):
            c = ws.cell(row=er, column=ci+1, value=v)
            c.font  = BF if ci == 0 else NF
            c.fill  = fill
            c.alignment = align
            c.border = BRD
            if fmt:
                c.number_format = fmt

    # Widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 6
    for col in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 11
    ws.column_dimensions["H"].width = 8
    ws.column_dimensions["I"].width = 22
    ws.column_dimensions["J"].width = 35
    ws.column_dimensions["K"].width = 45
    ws.column_dimensions["L"].width = 80

    ws.row_dimensions[1].height = 22
    for r in range(2, len(all_rows) + 2):
        ws.row_dimensions[r].height = 50
    ws.freeze_panes = "C2"


def _mode_fill(mode: str) -> PatternFill:
    return {"isolation": MODE_ISO, "optimization": MODE_OPT,
            "learning": MODE_LRN}.get(mode, MODE_NOD)


def _action_fill(action: str) -> PatternFill:
    a = (action or "").lower()
    if "increase" in a:   return ACT_INC
    if "reduce" in a:     return ACT_RED
    return ACT_KEEP


def _mode_label(mode: str) -> str:
    return {"isolation":   "🔴 ISOLATION",
            "optimization":"🟢 OPTIMIZATION",
            "learning":    "🚼 LEARNING",
            "no_data":     "⚫ NO DATA"}.get(mode, mode.upper() if mode else "—")


def _make_algo_sheet(ws, rows: list[CampaignResult]):
    """
    Placement Algorithm sheet.
    Layout (20 columns):
      A: Campaign | B: Mode | C: Base Bid Δ% | D: Score | E: Overall Reasoning
      F-J : Top of Search   (Current% | ROAS | Confidence | Rec% | Action)
      K-O : Rest of Search  (same 5)
      P-T : Product Pages   (same 5)
    """
    PL_HDRS  = ["Current %", "ROAS", "Confidence", "Rec %", "Action"]
    PL_FILLS = [H_ALGO_T, H_ALGO_R, H_ALGO_P]
    PL_NAMES = ["Top of Search", "Rest of Search", "Product Pages"]
    PL_KEYS  = ["Top of Search", "Rest of Search", "Product Pages"]

    # ── Row 1: group headers ──────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    _cell(ws, 1, 1, "Campaign & Algorithm", WF, H_ALGO, C)
    for i, (name, fill) in enumerate(zip(PL_NAMES, PL_FILLS)):
        start = 6 + i * 5
        end   = start + 4
        ws.merge_cells(f"{get_column_letter(start)}1:{get_column_letter(end)}1")
        _cell(ws, 1, start, name, WF, fill, C)

    # ── Row 2: column headers ─────────────────────────────────────────────
    info_hdrs = ["Campaign", "Mode", "Base Bid Change", "Score (1-10)", "Algorithm Reasoning"]
    for ci, h in enumerate(info_hdrs):
        _cell(ws, 2, ci + 1, h, WF, H_ALGO, C)
    for i, fill in enumerate(PL_FILLS):
        for j, h in enumerate(PL_HDRS):
            _cell(ws, 2, 6 + i * 5 + j, h, WF, fill, C)

    # ── Data rows ─────────────────────────────────────────────────────────
    for ri, r in enumerate(rows):
        er  = ri + 3
        bg  = ALT if ri % 2 == 0 else WHITE
        algo = r.placement_algorithm or {}
        mode = algo.get("mode", "")
        mf   = _mode_fill(mode)
        base_change = algo.get("base_bid_change_pct", 0)
        score       = algo.get("score", 0)
        reasoning   = algo.get("reasoning", "")
        placements  = {p["label"]: p for p in algo.get("placements", [])}

        _cell(ws, er, 1, r.campaign,              BF, bg,  L)
        _cell(ws, er, 2, _mode_label(mode),        NF, mf,  C)

        # Base bid change cell — red text if negative
        bc_val  = f"{'−' if base_change < 0 else ''}{abs(base_change)}%" if base_change != 0 else "No change"
        bc_font = Font(name="Arial", size=9, bold=True,
                       color="C00000" if base_change < 0 else "375623")
        c = ws.cell(row=er, column=3, value=bc_val)
        c.font = bc_font; c.fill = mf; c.alignment = C; c.border = BRD

        _cell(ws, er, 4, score if score else "—",  Font(name="Arial",bold=True,size=11),
              _score_fill(score * 10) if score else MODE_NOD, C)
        _cell(ws, er, 5, reasoning,                NF, bg, L)

        # Per-placement columns
        for i, pl_label in enumerate(PL_KEYS):
            p = placements.get(pl_label)
            base_col = 6 + i * 5
            if p:
                current_pct = f"{int(round(p.get('current_adj', 0) * 100))}%"
                rec_pct     = f"{p.get('recommended_multiplier', 0)}%"
                roas_val    = p.get('roas', 0)
                conf_val    = p.get('confidence', 0)
                action      = p.get('recommended_action', '—')
                af          = _action_fill(action)

                _cell(ws, er, base_col,     current_pct,             NF, bg, C)
                _cell(ws, er, base_col + 1, roas_val,                NF, bg, C, "0.00")
                _cell(ws, er, base_col + 2, f"{conf_val:.0%}",       NF, bg, C)
                _cell(ws, er, base_col + 3, rec_pct,
                      Font(name="Arial", bold=True, size=9), af, C)
                _cell(ws, er, base_col + 4, action,                  NF, af, C)
            else:
                for j in range(5):
                    _cell(ws, er, base_col + j, "—", NF, bg, C)

    # ── Column widths ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 55
    for col in range(6, 21):
        ltr = get_column_letter(col)
        ws.column_dimensions[ltr].width = 13 if (col - 6) % 5 == 4 else 11

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 22
    for r in range(3, len(rows) + 3):
        ws.row_dimensions[r].height = 55
    ws.freeze_panes = "C3"


def build_excel(results: list[CampaignResult]) -> bytes:
    """
    Build the full 3-sheet Excel and return as bytes (for Streamlit download).
    """
    sp_rows = [r for r in results if r.ad_type == "SP"]
    sb_rows = [r for r in results if r.ad_type == "SB"]

    wb = Workbook()

    ws_sp = wb.active
    ws_sp.title = "Sponsored Products"
    _make_detail_sheet(ws_sp, sp_rows)

    ws_sb = wb.create_sheet("Sponsored Brands")
    _make_detail_sheet(ws_sb, sb_rows)

    ws_sum = wb.create_sheet("Top of Search — SP vs SB")
    _make_summary_sheet(ws_sum, sp_rows, sb_rows)

    ws_algo = wb.create_sheet("📍 Placement Algorithm")
    _make_algo_sheet(ws_algo, results)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
